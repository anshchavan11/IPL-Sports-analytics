import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.worksheet.table import Table, TableStyleInfo

DATA = "/home/claude/ipl-analytics-project/data/"

matches = pd.read_csv(DATA+"matches.csv")
team_summary = pd.read_csv(DATA+"team_summary.csv")
batters = pd.read_csv(DATA+"batter_summary.csv").head(50)
bowlers = pd.read_csv(DATA+"bowler_summary.csv").head(50)
season_summary = pd.read_csv(DATA+"season_summary.csv")
deliveries_sample = pd.read_csv(DATA+"deliveries.csv").head(3000)

wb = Workbook()

NAVY = "1E2761"
ICE = "CADCFC"
WHITE = "FFFFFF"
GOLD = "D4A017"
LIGHT_GRAY = "F2F2F2"

FONT_NAME = "Calibri"
header_font = Font(name=FONT_NAME, bold=True, color=WHITE, size=11)
header_fill = PatternFill("solid", fgColor=NAVY)
title_font = Font(name=FONT_NAME, bold=True, color=NAVY, size=16)
subtitle_font = Font(name=FONT_NAME, italic=True, color="555555", size=10)
body_font = Font(name=FONT_NAME, size=10)
thin_border = Border(*(Side(style='thin', color='D9D9D9'),)*4)

def style_header_row(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

def autosize(ws, ncols, min_w=10, max_w=42):
    for c in range(1, ncols+1):
        col = get_column_letter(c)
        max_len = min_w
        for cell in ws[col]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col].width = min(max_len + 2, max_w)

def write_df(ws, df, start_row=1, start_col=1, table_name=None, style="TableStyleMedium9"):
    for j, col in enumerate(df.columns):
        ws.cell(row=start_row, column=start_col+j, value=col)
    for i, row in enumerate(df.itertuples(index=False), start=1):
        for j, val in enumerate(row):
            ws.cell(row=start_row+i, column=start_col+j, value=val)
    style_header_row(ws, start_row, len(df.columns))
    end_row = start_row + len(df)
    end_col = start_col + len(df.columns) - 1
    if table_name:
        ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        tab = Table(displayName=table_name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name=style, showRowStripes=True)
        ws.add_table(tab)
    return end_row, end_col

# ============================================================
# Sheet 1: README / Cover
# ============================================================
ws = wb.active
ws.title = "README"
ws["B2"] = "IPL Sports Analytics — Excel Workbook"
ws["B2"].font = Font(name=FONT_NAME, bold=True, size=20, color=NAVY)
ws["B3"] = "2016–2024 seasons | SQL + Excel + Dashboard portfolio project"
ws["B3"].font = subtitle_font

notes = [
    "", "How this workbook is organized:",
    "1. Matches_Raw — one row per match (606 matches, 2016-2024), source table",
    "2. Deliveries_Sample — ball-by-ball data sample (full 144K-row set lives in /data and the SQL database)",
    "3. Lookups — XLOOKUP / INDEX-MATCH / VLOOKUP demos pulling match details by Match ID",
    "4. Team_Pivot — SUMIFS/COUNTIFS pivot-style summary: win %, matches played per team",
    "5. Player_Pivot — Top run scorers & wicket takers with strike rate / economy",
    "6. Season_Pivot — Orange Cap / Purple Cap holder per season",
    "7. Dashboard — KPI cards + charts pulling live from the sheets above",
    "", "Data source: schema matches the public Kaggle dataset",
    "'IPL Complete Dataset (2008-2020)' by patrickb1912 (matches.csv + deliveries.csv).",
    "This workbook uses a realistic simulated dataset with the identical schema —",
    "see README.md in the project root for details and how to swap in the real Kaggle CSVs."
]
for i, line in enumerate(notes, start=5):
    cell = ws.cell(row=i, column=2, value=line)
    if line.endswith(":") or line.startswith(tuple("1234567")):
        cell.font = Font(name=FONT_NAME, bold=True, size=11) if line.endswith(":") else Font(name=FONT_NAME, size=10)
    else:
        cell.font = Font(name=FONT_NAME, size=10)
ws.column_dimensions['B'].width = 95
ws.sheet_view.showGridLines = False

# ============================================================
# Sheet 2: Matches_Raw
# ============================================================
ws = wb.create_sheet("Matches_Raw")
matches_x = matches.rename(columns={"id": "match_id"})
write_df(ws, matches_x, table_name="MatchesRaw")
autosize(ws, len(matches_x.columns))
ws.freeze_panes = "A2"

# ============================================================
# Sheet 3: Deliveries_Sample
# ============================================================
ws = wb.create_sheet("Deliveries_Sample")
write_df(ws, deliveries_sample, table_name="DeliveriesSample")
autosize(ws, len(deliveries_sample.columns))
ws.freeze_panes = "A2"

# ============================================================
# Sheet 4: Lookups (VLOOKUP / XLOOKUP / INDEX-MATCH demo)
# ============================================================
ws = wb.create_sheet("Lookups")
ws["B2"] = "Lookup Formula Demonstrations"
ws["B2"].font = title_font
ws["B3"] = "Enter a Match ID in cell C6 to pull match details using 3 different lookup methods"
ws["B3"].font = subtitle_font

ws["B6"] = "Enter Match ID:"
ws["B6"].font = Font(name=FONT_NAME, bold=True)
ws["C6"] = 101
ws["C6"].font = Font(name=FONT_NAME, bold=True, color="B85042")
ws["C6"].fill = PatternFill("solid", fgColor="FFFF00")

labels = ["Method", "Winner", "Venue", "City", "Toss Winner", "Player of Match"]
for j, lab in enumerate(labels):
    ws.cell(row=8, column=2+j, value=lab).font = header_font
    ws.cell(row=8, column=2+j).fill = header_fill

# VLOOKUP row
ws["B9"] = "VLOOKUP"
ws["C9"] = "=VLOOKUP($C$6,Matches_Raw!A:N,10,FALSE)"
ws["D9"] = "=VLOOKUP($C$6,Matches_Raw!A:N,6,FALSE)"
ws["E9"] = "=VLOOKUP($C$6,Matches_Raw!A:N,7,FALSE)"
ws["F9"] = "=VLOOKUP($C$6,Matches_Raw!A:N,8,FALSE)"
ws["G9"] = "=VLOOKUP($C$6,Matches_Raw!A:N,13,FALSE)"

# INDEX-MATCH row
ws["B10"] = "INDEX-MATCH"
ws["C10"] = "=INDEX(Matches_Raw!J:J,MATCH($C$6,Matches_Raw!A:A,0))"
ws["D10"] = "=INDEX(Matches_Raw!F:F,MATCH($C$6,Matches_Raw!A:A,0))"
ws["E10"] = "=INDEX(Matches_Raw!G:G,MATCH($C$6,Matches_Raw!A:A,0))"
ws["F10"] = "=INDEX(Matches_Raw!H:H,MATCH($C$6,Matches_Raw!A:A,0))"
ws["G10"] = "=INDEX(Matches_Raw!M:M,MATCH($C$6,Matches_Raw!A:A,0))"

# XLOOKUP row
ws["B11"] = "XLOOKUP"
ws["C11"] = "=XLOOKUP($C$6,Matches_Raw!A:A,Matches_Raw!J:J)"
ws["D11"] = "=XLOOKUP($C$6,Matches_Raw!A:A,Matches_Raw!F:F)"
ws["E11"] = "=XLOOKUP($C$6,Matches_Raw!A:A,Matches_Raw!G:G)"
ws["F11"] = "=XLOOKUP($C$6,Matches_Raw!A:A,Matches_Raw!H:H)"
ws["G11"] = "=XLOOKUP($C$6,Matches_Raw!A:A,Matches_Raw!M:M)"

for r in (9, 10, 11):
    ws.cell(row=r, column=2).font = Font(name=FONT_NAME, bold=True, color=NAVY)
    for c in range(3, 8):
        ws.cell(row=r, column=c).border = thin_border

ws["B14"] = "Team Short-Name Lookup Table (used by charts across this workbook)"
ws["B14"].font = Font(name=FONT_NAME, bold=True)
team_short = pd.DataFrame({
    "team": ["Mumbai Indians","Chennai Super Kings","Royal Challengers Bangalore","Kolkata Knight Riders",
             "Delhi Capitals","Punjab Kings","Rajasthan Royals","Sunrisers Hyderabad","Gujarat Titans","Lucknow Super Giants"],
    "short": ["MI","CSK","RCB","KKR","DC","PBKS","RR","SRH","GT","LSG"]
})
write_df(ws, team_short, start_row=15, start_col=2, table_name="TeamShort")
autosize(ws, 8)
ws.column_dimensions['B'].width = 16

# ============================================================
# Sheet 5: Team_Pivot (SUMIFS / COUNTIFS pivot-style summary)
# ============================================================
ws = wb.create_sheet("Team_Pivot")
ws["B2"] = "Team Performance Summary (Pivot-Style, built with SUMIFS/COUNTIFS)"
ws["B2"].font = title_font

teams = team_summary['team'].tolist()
headers = ["Team", "Matches Played", "Matches Won", "Win %"]
for j, h in enumerate(headers):
    ws.cell(row=4, column=2+j, value=h)
style_header_row(ws, 4, 4)

for i, team in enumerate(teams, start=5):
    ws.cell(row=i, column=2, value=team)
    ws.cell(row=i, column=3,
        value=f'=COUNTIF(Matches_Raw!D:D,B{i})+COUNTIF(Matches_Raw!E:E,B{i})')
    ws.cell(row=i, column=4,
        value=f'=COUNTIF(Matches_Raw!J:J,B{i})')
    ws.cell(row=i, column=5, value=f'=ROUND(D{i}/C{i}*100,2)')
    for c in range(2, 6):
        ws.cell(row=i, column=c).border = thin_border

last = 4 + len(teams)
# Data bars on win %
rule = DataBarRule(start_type='min', end_type='max', color="638EC6")
ws.conditional_formatting.add(f"E5:E{last}", rule)
autosize(ws, 6)
ws.column_dimensions['B'].width = 28

# ============================================================
# Sheet 6: Player_Pivot
# ============================================================
ws = wb.create_sheet("Player_Pivot")
ws["B2"] = "Player Performance — Top Run Scorers & Wicket Takers"
ws["B2"].font = title_font

ws["B4"] = "Top 15 Run Scorers"
ws["B4"].font = Font(name=FONT_NAME, bold=True, size=12, color=NAVY)
bat_top = batters[['batter','total_runs','innings','strike_rate','sixes','fours']].head(15)
bat_top.columns = ['Batter','Total Runs','Innings','Strike Rate','6s','4s']
end_r, _ = write_df(ws, bat_top, start_row=5, start_col=2, table_name="TopBatters")
rule = ColorScaleRule(start_type='min', start_color='F8696B', end_type='max', end_color='63BE7B')
ws.conditional_formatting.add(f"C6:C{end_r}", rule)

start2 = end_r + 3
ws.cell(row=start2-1, column=2, value="Top 15 Wicket Takers").font = Font(name=FONT_NAME, bold=True, size=12, color=NAVY)
bowl_top = bowlers[['bowler','wickets','matches','economy','avg_wkts_per_match']].head(15)
bowl_top.columns = ['Bowler','Wickets','Matches','Economy','Avg Wkts/Match']
end_r2, _ = write_df(ws, bowl_top, start_row=start2, start_col=2, table_name="TopBowlers")
rule2 = ColorScaleRule(start_type='min', start_color='F8696B', end_type='max', end_color='63BE7B')
ws.conditional_formatting.add(f"C{start2+1}:C{end_r2}", rule2)

autosize(ws, 7)
ws.column_dimensions['B'].width = 24

# ============================================================
# Sheet 7: Season_Pivot
# ============================================================
ws = wb.create_sheet("Season_Pivot")
ws["B2"] = "Season-by-Season Summary — Orange Cap & Purple Cap"
ws["B2"].font = title_font
season_x = season_summary[['season','matches_played','orange_cap','runs','purple_cap','wickets']]
season_x.columns = ['Season','Matches Played','Orange Cap (Most Runs)','Runs','Purple Cap (Most Wickets)','Wickets']
write_df(ws, season_x, start_row=4, start_col=2, table_name="SeasonPivot")
autosize(ws, 7)
ws.column_dimensions['B'].width = 12

wb.save("/home/claude/ipl-analytics-project/excel/IPL_Analytics.xlsx")
print("Workbook saved (part 1 - data/pivots)")
