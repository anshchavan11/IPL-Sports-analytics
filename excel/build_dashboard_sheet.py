from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

path = "/home/claude/ipl-analytics-project/excel/IPL_Analytics.xlsx"
wb = load_workbook(path)

NAVY = "1E2761"
ICE = "CADCFC"
WHITE = "FFFFFF"
GOLD = "D4A017"
FONT_NAME = "Calibri"

ws = wb.create_sheet("Dashboard", 0)  # make it the first visible sheet after README (we'll reorder)
ws.sheet_view.showGridLines = False

ws["B2"] = "IPL Analytics Dashboard"
ws["B2"].font = Font(name=FONT_NAME, bold=True, size=22, color=NAVY)
ws["B3"] = "2016 – 2024  |  9 Seasons  |  606 Matches"
ws["B3"].font = Font(name=FONT_NAME, italic=True, size=11, color="666666")

# ---- KPI Cards ----
kpis = [
    ("Total Matches", "=COUNTA(Matches_Raw!A2:A607)", "B"),
    ("Total Seasons", "=COUNTA(Season_Pivot!C5:C13)", "F"),
    ("Most Successful Team", "=INDEX(Team_Pivot!C5:C14,MATCH(MAX(Team_Pivot!F5:F14),Team_Pivot!F5:F14,0))", "J"),
    ("Highest Run Scorer", "=Player_Pivot!C6", "N"),
]
col_map = {"B":2, "F":6, "J":10, "N":14}
for label, formula, col in kpis:
    c = col_map[col]
    ws.cell(row=6, column=c, value=label).font = Font(name=FONT_NAME, size=10, color="666666")
    val_cell = ws.cell(row=7, column=c, value=formula)
    val_cell.font = Font(name=FONT_NAME, bold=True, size=16, color=NAVY)
    for r in (6,7):
        cell = ws.cell(row=r, column=c)
        cell.fill = PatternFill("solid", fgColor="F2F2F2")
    ws.merge_cells(start_row=6, start_column=c, end_row=6, end_column=c+2)
    ws.merge_cells(start_row=7, start_column=c, end_row=7, end_column=c+2)
    for cc in range(c, c+3):
        ws.cell(row=6, column=cc).fill = PatternFill("solid", fgColor="F2F2F2")
        ws.cell(row=7, column=cc).fill = PatternFill("solid", fgColor="F2F2F2")

# Fix win% card to reference correct column (E = Win %, not F)
ws.cell(row=7, column=10, value="=INDEX(Team_Pivot!C5:C14,MATCH(MAX(Team_Pivot!E5:E14),Team_Pivot!E5:E14,0))")

# ============================================================
# Chart 1: Team Wins - Bar Chart
# ============================================================
bar = BarChart()
bar.title = "Total Wins by Team (2016-2024)"
bar.style = 10
bar.y_axis.title = "Wins"
bar.x_axis.title = "Team"
data = Reference(wb["Team_Pivot"], min_col=4, min_row=4, max_row=14)
cats = Reference(wb["Team_Pivot"], min_col=2, min_row=5, max_row=14)
bar.add_data(data, titles_from_data=True)
bar.set_categories(cats)
bar.height = 9
bar.width = 18
bar.legend = None
ws.add_chart(bar, "B10")

# ============================================================
# Chart 2: Runs by Season - Line Chart
# ============================================================
line = LineChart()
line.title = "Orange Cap Runs by Season"
line.y_axis.title = "Runs"
line.x_axis.title = "Season"
data = Reference(wb["Season_Pivot"], min_col=5, min_row=4, max_row=13)
cats = Reference(wb["Season_Pivot"], min_col=2, min_row=5, max_row=13)
line.add_data(data, titles_from_data=True)
line.set_categories(cats)
line.height = 9
line.width = 18
ws.add_chart(line, "J10")

# ============================================================
# Chart 3: Toss Decision Split - Pie Chart
# ============================================================
ws["B28"] = "Toss Decision"
ws["B28"].font = Font(name=FONT_NAME, bold=True, size=11, color=NAVY)
ws["B29"] = "Bat First"
ws["C29"] = '=COUNTIF(Matches_Raw!H:H,"bat")'
ws["B30"] = "Field First"
ws["C30"] = '=COUNTIF(Matches_Raw!H:H,"field")'

pie = PieChart()
pie.title = "Toss Decision Split"
data = Reference(ws, min_col=3, min_row=29, max_row=30)
cats = Reference(ws, min_col=2, min_row=29, max_row=30)
pie.add_data(data)
pie.set_categories(cats)
pie.height = 8
pie.width = 10
pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = True
ws.add_chart(pie, "B32")

# ============================================================
# Chart 4: Top 10 Batters - Bar Chart (horizontal feel via bar)
# ============================================================
bar2 = BarChart()
bar2.type = "bar"
bar2.title = "Top 10 Run Scorers"
bar2.y_axis.title = "Player"
bar2.x_axis.title = "Runs"
data = Reference(wb["Player_Pivot"], min_col=3, min_row=5, max_row=15)
cats = Reference(wb["Player_Pivot"], min_col=2, min_row=6, max_row=15)
bar2.add_data(data, titles_from_data=True)
bar2.set_categories(cats)
bar2.height = 9
bar2.width = 18
bar2.legend = None
ws.add_chart(bar2, "J32")

for c in range(2, 17):
    ws.column_dimensions[get_column_letter(c)].width = 11

wb.save(path)

# Reorder sheets: README, Dashboard, then the rest
order = ["README", "Dashboard", "Matches_Raw", "Deliveries_Sample", "Lookups", "Team_Pivot", "Player_Pivot", "Season_Pivot"]
wb2 = load_workbook(path)
wb2._sheets = [wb2[name] for name in order]
for name in order:
    wb2[name].sheet_view.tabSelected = False
wb2.active = 1
wb2.save(path)
print("Dashboard added and sheets reordered")
